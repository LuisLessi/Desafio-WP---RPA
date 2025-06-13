#!/usr/bin/env python
# coding: utf-8

# In[13]:


import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Reference


# In[ ]:


print("🚀 Inicializando o WebDriver e acessando a página...")
driver = webdriver.Chrome()
driver.maximize_window() 
url = "https://finance.yahoo.com/quote/%5EBVSP/components/?p=%5EBVSP"
driver.get(url)
time.sleep(5)
print("✅ Página carregada!")

# Encontrar a tabela
print("🔍 Buscando a tabela de dados...")
table = driver.find_element(By.XPATH, "//table[@data-testid='table-container']")
rows = table.find_elements(By.TAG_NAME, "tr")
print("✅ Tabela encontrada!")

# Extrair os dados
print("📊 Extraindo informações da tabela...")
data = []
for row in rows[1:]:  # Ignorar cabeçalho
    cols = row.find_elements(By.TAG_NAME, "td")
    if len(cols) >= 6:
        symbol = cols[0].text.strip()
        company = cols[1].text.strip()
        last_price = cols[2].text.strip()
        change = cols[3].text.strip()
        percent_change = cols[4].text.strip()
        volume = cols[5].text.strip()
        data.append([symbol, company, last_price, change, percent_change, volume])
print("✅ Dados extraídos com sucesso!")

# Criar DataFrame
df = pd.DataFrame(data, columns=["Código", "Empresa", "Último Preço", "Variação", "% Variação", "Volume"])

# Criar planilha com três abas
print("📁 Criando o arquivo Excel...")
wb = Workbook()

# Aba 1: Dados Brutos
ws1 = wb.active
ws1.title = "Dados Brutos"
ws1.append(df.columns.tolist())

# Aplicar formatação ao cabeçalho
header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
header_font = Font(bold=True)

for col in range(1, len(df.columns) + 1):
    cell = ws1.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font

for row in df.itertuples(index=False):
    ws1.append(list(row))
print("✅ Aba 'Dados Brutos' criada!")

# Aba 2: Dados Tratados
print("🛠 Processando e categorizando os dados...")
ws2 = wb.create_sheet(title="Dados Tratados")

df_cleaned = df[df["Último Preço"] != "--"].copy()
df_cleaned.loc[:, "% Variação"] = df_cleaned["% Variação"].str.replace("%", "").astype(float)
df_cleaned = df_cleaned.sort_values("% Variação", ascending=False)
df_cleaned.loc[:, "Categoria"] = df_cleaned["% Variação"].apply(lambda x: "Alta" if x > 2 else "Queda" if x < -2 else "Estável")

ws2.append(df_cleaned.columns.tolist())

# Aplicar formatação ao cabeçalho da aba 'Dados Tratados'
for col in range(1, len(df_cleaned.columns) + 1):
    cell = ws2.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font

for row in df_cleaned.itertuples(index=False):
    ws2.append(list(row))
print("✅ Aba 'Dados Tratados' criada!")

# Aba 3: Análises
print("📈 Gerando estatísticas e gráfico de distribuição...")
ws3 = wb.create_sheet(title="Análises")

media_precos = df_cleaned["Último Preço"].astype(float).mean()
maior_var = df_cleaned.loc[df_cleaned["% Variação"].idxmax()]
menor_var = df_cleaned.loc[df_cleaned["% Variação"].idxmin()]
categorias = df_cleaned["Categoria"].value_counts()

ws3.append(["Média dos Preços", media_precos])
ws3.append(["Empresa com Maior Variação", maior_var["Empresa"], maior_var["% Variação"]])
ws3.append(["Empresa com Menor Variação", menor_var["Empresa"], menor_var["% Variação"]])
ws3.append(["Distribuição das Categorias"])

for cat, count in categorias.items():
    ws3.append([cat, count])

# Criar gráfico de barras no Excel
chart = BarChart()
chart.title = "Distribuição das Categorias"
chart.x_axis.title = "Categoria"
chart.y_axis.title = "Número de Empresas"

data_ref = Reference(ws3, min_col=2, min_row=5, max_row=7)  # Inclui todas as categorias
cats_ref = Reference(ws3, min_col=1, min_row=5, max_row=7)

chart.add_data(data_ref, titles_from_data=False)
chart.set_categories(cats_ref)
ws3.add_chart(chart, "E10")  # Posiciona o gráfico na célula E10
print("✅ Gráfico de distribuição adicionado!")

# Salvar planilha
wb.save("dados_bovespa.xlsx")
driver.quit()
print("✅ Processo concluído! Planilha 'dados_bovespa.xlsx' gerada com sucesso! 🚀")


# In[ ]:




